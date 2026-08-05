"""Parsing, validation, duplicate detection, and confirmed statement imports."""

import csv
import io
import logging
import re
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from app.ai.llm_gateway import LLMGateway
from app.core.exceptions import BadRequestException, ConflictException, NotFoundException
from app.models.statement import Statement, StatementStatus
from app.models.transaction import TransactionType
from app.repositories.account_repository import AccountRepository
from app.repositories.category_repository import CategoryRepository
from app.repositories.statement_repository import StatementRepository
from app.repositories.transaction_repository import TransactionRepository
from app.schemas.statement import StatementImportResponse, StatementPreviewTransaction, StatementUploadPreview
from app.schemas.transaction import TransactionCreate
from app.services.transaction_service import TransactionService

logger = logging.getLogger(__name__)
MAX_FILE_SIZE = 10 * 1024 * 1024
ALLOWED_TYPES = {".csv": {"text/csv", "application/csv", "text/plain"}, ".xlsx": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/octet-stream"}}
CANONICAL_CATEGORIES = {"food": "Food", "shopping": "Shopping", "transport": "Transport", "bills": "Bills", "entertainment": "Entertainment", "healthcare": "Healthcare", "education": "Education", "travel": "Travel", "salary": "Salary", "investment": "Investment", "miscellaneous": "Miscellaneous", "uncategorized": "Uncategorized"}


class StatementService:
    def __init__(self, statements: StatementRepository, transactions: TransactionRepository, accounts: AccountRepository, categories: CategoryRepository, transaction_service: TransactionService) -> None:
        self._statements, self._transactions, self._accounts = statements, transactions, accounts
        self._categories, self._transaction_service = categories, transaction_service

    def preview_upload(self, user_id: int, account_id: int, filename: str, content_type: str | None, data: bytes) -> StatementUploadPreview:
        logger.info("Statement upload started user_id=%s", user_id)
        suffix = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if suffix not in ALLOWED_TYPES or (content_type and content_type.split(";", 1)[0].lower() not in ALLOWED_TYPES[suffix]):
            raise BadRequestException("Only CSV and XLSX statement files are supported.")
        if not data:
            raise BadRequestException("The uploaded file is empty.")
        if len(data) > MAX_FILE_SIZE:
            raise BadRequestException("The uploaded file exceeds the 10 MB limit.")
        if self._accounts.get_by_id_and_user_id(account_id, user_id) is None:
            raise NotFoundException("Account not found.")
        rows = self._read_rows(suffix, data)
        if not rows:
            raise BadRequestException("The statement contains no data rows.")
        normalized = self._normalize_rows(rows)
        if normalized is None:
            raise BadRequestException("Required statement columns are missing.")
        visible_categories = {category.name.casefold(): category for category in self._categories.list_visible(user_id)}
        previews, warnings, duplicate_count = [], [], 0
        seen: set[tuple] = set()
        for index, raw in enumerate(normalized, start=2):
            parsed = self._parse_row(index, raw, visible_categories)
            if parsed.valid:
                key = (parsed.date.isoformat(), str(parsed.amount), (parsed.merchant or parsed.description or "").casefold(), parsed.transaction_type.value)
                parsed.duplicate = key in seen or self._duplicate_exists(user_id, parsed)
                seen.add(key)
                duplicate_count += int(parsed.duplicate)
            else:
                warnings.append(f"Row {index}: {parsed.error}")
            previews.append(parsed)
        valid = sum(item.valid for item in previews)
        dates = [item.date for item in previews if item.date]
        representative = dates[0] if dates else date.today()
        stored = [item.model_dump(mode="json") for item in previews]
        statement = self._statements.create(Statement(user_id=user_id, account_id=account_id, filename=filename[:255], bank_name=None, statement_month=representative.month, statement_year=representative.year, file_type=suffix[1:], total_transactions=valid, imported_transactions=0, status=StatementStatus.PREVIEWED, preview_data=stored))
        logger.info("Statement rows parsed user_id=%s rows=%s duplicates=%s", user_id, len(rows), duplicate_count)
        return StatementUploadPreview(statement_id=statement.id, total_rows=len(rows), valid_rows=valid, invalid_rows=len(rows)-valid, duplicate_rows=duplicate_count, preview_transactions=previews, warnings=warnings)

    def import_statement(self, user_id: int, statement_id: int) -> StatementImportResponse:
        statement = self._statements.get_by_id_and_user_id(statement_id, user_id)
        if statement is None:
            raise NotFoundException("Statement not found.")
        if statement.status == StatementStatus.IMPORTED:
            raise ConflictException("This statement has already been imported.")
        imported = skipped = 0
        try:
            for raw in statement.preview_data:
                item = StatementPreviewTransaction.model_validate(raw)
                if not item.valid or item.duplicate:
                    skipped += int(item.valid and item.duplicate)
                    continue
                if self._duplicate_exists(user_id, item):
                    skipped += 1
                    continue
                category_id = self._category_id(user_id, item.category)
                self._transaction_service.create_transaction(user_id, TransactionCreate(account_id=statement.account_id, category_id=category_id, transaction_type=item.transaction_type, title=(item.merchant or item.description or "Imported transaction")[:255], description=item.description, amount=item.amount, transaction_date=item.date, merchant=item.merchant, tags=[]))
                imported += 1
        except Exception:
            statement.status = StatementStatus.FAILED
            self._statements.save(statement)
            logger.exception("Statement import failed user_id=%s statement_id=%s", user_id, statement_id)
            raise
        # The source file is never stored and the transient preview is no longer needed.
        statement.preview_data = []
        statement.imported_transactions, statement.status = imported, StatementStatus.IMPORTED
        self._statements.save(statement)
        logger.info("Statement import completed user_id=%s statement_id=%s imported=%s skipped_duplicates=%s", user_id, statement_id, imported, skipped)
        return StatementImportResponse(statement_id=statement.id, status=statement.status, imported_transactions=imported, skipped_duplicates=skipped)

    def _read_rows(self, suffix: str, data: bytes) -> list[dict[str, object]]:
        try:
            if suffix == ".csv":
                text = data.decode("utf-8-sig")
                return list(csv.DictReader(io.StringIO(text)))
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
            if not values: return []
            headers = [str(value).strip() if value is not None else "" for value in values[0]]
            return [dict(zip(headers, row)) for row in values[1:] if any(value is not None and str(value).strip() for value in row)]
        except (UnicodeDecodeError, csv.Error, ValueError, OSError, Exception) as exc:
            logger.warning("Statement parsing failed error=%s", type(exc).__name__)
            raise BadRequestException("The uploaded statement file is invalid or corrupted.") from None

    def _normalize_rows(self, rows: list[dict[str, object]]) -> list[dict[str, object]] | None:
        aliases = {"date": {"date", "transaction date"}, "description": {"description", "narration", "details"}, "merchant": {"merchant"}, "debit": {"debit"}, "credit": {"credit"}, "amount": {"amount"}, "category": {"category"}}
        mapping = {}
        for header in rows[0]:
            cleaned = re.sub(r"\s+", " ", str(header).strip().casefold())
            for target, names in aliases.items():
                if cleaned in names: mapping[target] = header
        available = set(mapping)
        if "date" not in mapping or not ({"description", "merchant"} & available) or not ({"amount", "debit", "credit"} & available): return None
        return [{target: row.get(source) for target, source in mapping.items()} for row in rows]

    def _parse_row(self, row_number: int, row: dict[str, object], categories: dict) -> StatementPreviewTransaction:
        try:
            transaction_date = self._parse_date(row.get("date"))
            description = self._text(row.get("description"))
            merchant = self._text(row.get("merchant")) or description
            if not merchant: raise ValueError("merchant or description is required")
            amount, kind = self._amount_and_type(row)
            provided = self._text(row.get("category"))
            category = self._resolve_category(provided, merchant, categories)
            return StatementPreviewTransaction(row_number=row_number, date=transaction_date, merchant=merchant[:255], description=description[:10000] if description else None, amount=amount, transaction_type=kind, category=category, valid=True)
        except (ValueError, InvalidOperation) as exc:
            return StatementPreviewTransaction(row_number=row_number, valid=False, error=str(exc))

    @staticmethod
    def _text(value: object) -> str | None:
        value = str(value).strip() if value is not None else ""
        return value or None

    @staticmethod
    def _parse_date(value: object) -> date:
        if isinstance(value, date): return value
        text = str(value).strip()
        from datetime import datetime
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
            try: return datetime.strptime(text, fmt).date()
            except ValueError: pass
        raise ValueError("invalid date")

    def _amount_and_type(self, row: dict[str, object]) -> tuple[Decimal, TransactionType]:
        debit, credit, amount = (self._decimal(row.get(key)) for key in ("debit", "credit", "amount"))
        if debit is not None and debit > 0: return debit, TransactionType.EXPENSE
        if credit is not None and credit > 0: return credit, TransactionType.INCOME
        if amount is None or amount == 0: raise ValueError("amount is required")
        return abs(amount), TransactionType.INCOME if amount > 0 else TransactionType.EXPENSE

    @staticmethod
    def _decimal(value: object) -> Decimal | None:
        if value is None or str(value).strip() == "": return None
        cleaned = re.sub(r"[^0-9.\-]", "", str(value).replace(",", ""))
        return Decimal(cleaned)

    def _resolve_category(self, supplied: str | None, merchant: str, categories: dict) -> str | None:
        if supplied and supplied.casefold() in categories: return categories[supplied.casefold()].name
        if supplied: return self._uncategorized(categories)
        heuristic = next((name for key, name in CANONICAL_CATEGORIES.items() if key in merchant.casefold()), None)
        if heuristic and heuristic.casefold() in categories: return categories[heuristic.casefold()].name
        try:
            answer = LLMGateway().generate(f"Classify this bank transaction into exactly one of: {', '.join(CANONICAL_CATEGORIES.values())}. Reply only with the category name. Transaction: {merchant}", temperature=0, max_tokens=10).content.strip().casefold()
            if answer in categories and answer in CANONICAL_CATEGORIES: return categories[answer].name
        except Exception:
            logger.info("Statement categorization unavailable; using Uncategorized")
        return self._uncategorized(categories)

    @staticmethod
    def _uncategorized(categories: dict) -> str | None:
        return categories["uncategorized"].name if "uncategorized" in categories else None

    def _category_id(self, user_id: int, category_name: str | None) -> int | None:
        if not category_name: return None
        return next((category.id for category in self._categories.list_visible(user_id) if category.name.casefold() == category_name.casefold()), None)

    def _duplicate_exists(self, user_id: int, item: StatementPreviewTransaction) -> bool:
        merchant = (item.merchant or item.description or "").casefold()
        return self._statements.transaction_duplicate_exists(
            user_id, item.date, item.amount, merchant, item.transaction_type
        )
